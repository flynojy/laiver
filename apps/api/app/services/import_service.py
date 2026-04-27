from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.enums import ImportSourceType, ImportStatus, MessageRole
from app.models.import_job import ImportJob, NormalizedMessage
from app.schemas.import_job import (
    ImportCommitRequest,
    ImportDetailResponse,
    ImportPreviewResponse,
    NormalizedMessageBase,
)
from app.utils.text import normalize_whitespace

WECHAT_SOURCE_FORMAT = "wechat_weflow_xlsx"
WECHAT_HEADER_ALIASES = {
    "sequence": {"\u5e8f\u53f7"},
    "time": {"\u65f6\u95f4", "\u53d1\u9001\u65f6\u95f4"},
    "sender": {"\u53d1\u9001\u8005\u8eab\u4efd", "\u53d1\u9001\u8005", "\u53d1\u4ef6\u4eba"},
    "message_type": {"\u6d88\u606f\u7c7b\u578b", "\u7c7b\u578b"},
    "content": {"\u5185\u5bb9", "\u6d88\u606f\u5185\u5bb9"},
}
WECHAT_TEXT_MESSAGE_TYPES = {"\u6587\u672c\u6d88\u606f"}
WECHAT_SYSTEM_MESSAGE_TYPES = {"\u7cfb\u7edf\u6d88\u606f"}
WECHAT_SKIP_PLACEHOLDERS = {
    "",
    "[\u5176\u4ed6\u6d88\u606f]",
    "[\u56fe\u7247]",
    "[\u8bed\u97f3]",
    "[\u89c6\u9891]",
    "[\u6587\u4ef6]",
    "[\u52a8\u753b\u8868\u60c5]",
    "[\u8868\u60c5]",
    "[emoji]",
}
WECHAT_SELF_SPEAKERS = {"\u6211", "me", "self"}


def _build_source_metadata(items: list[NormalizedMessageBase]) -> dict[str, Any]:
    speaker_stats: dict[str, dict[str, Any]] = {}
    message_types: set[str] = set()
    source_format = None
    conversation_owner = None
    export_tool = None
    export_version = None
    platform = None
    exported_at = None

    for item in items:
        speaker_entry = speaker_stats.setdefault(
            item.speaker,
            {"message_count": 0, "roles": [], "is_self": False},
        )
        speaker_entry["message_count"] += 1
        role_value = item.role.value
        if role_value not in speaker_entry["roles"]:
            speaker_entry["roles"].append(role_value)
        if item.metadata.get("is_self"):
            speaker_entry["is_self"] = True

        message_type = str(item.metadata.get("message_type") or "").strip()
        if message_type:
            message_types.add(message_type)
        source_format = source_format or item.metadata.get("source_format")
        conversation_owner = conversation_owner or item.metadata.get("conversation_owner")
        export_tool = export_tool or item.metadata.get("export_tool")
        export_version = export_version or item.metadata.get("export_version")
        platform = platform or item.metadata.get("platform")
        exported_at = exported_at or item.metadata.get("exported_at")

    return {
        "source_format": source_format,
        "conversation_owner": conversation_owner,
        "export_tool": export_tool,
        "export_version": export_version,
        "platform": platform,
        "exported_at": exported_at,
        "message_types": sorted(message_types),
        "speaker_stats": speaker_stats,
    }


def detect_source_type(file_name: str) -> ImportSourceType:
    suffix = file_name.rsplit(".", 1)[-1].lower()
    if suffix == "csv":
        return ImportSourceType.CSV
    if suffix == "json":
        return ImportSourceType.JSON
    if suffix == "xlsx":
        return ImportSourceType.XLSX
    return ImportSourceType.TXT


def infer_role(speaker: str) -> MessageRole:
    lowered = speaker.lower()
    if lowered in {"assistant", "ai", "bot"}:
        return MessageRole.ASSISTANT
    if lowered in {"system", "sys"}:
        return MessageRole.SYSTEM
    if lowered in {"tool", "function"}:
        return MessageRole.TOOL
    return MessageRole.USER


def parse_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    value = str(value).strip()
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return normalize_whitespace(str(value))


def _wechat_header_key(value: str) -> str | None:
    normalized = normalize_whitespace(value)
    for key, aliases in WECHAT_HEADER_ALIASES.items():
        if normalized in aliases:
            return key
    return None


def _parse_wechat_metadata(rows: list[tuple[Any, ...]]) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for row in rows:
        cells = [_cell_text(cell) for cell in row]
        index = 0
        while index < len(cells):
            label = cells[index]
            if not label:
                index += 1
                continue
            value_index = index + 1
            while value_index < len(cells) and not cells[value_index]:
                value_index += 1
            value = cells[value_index] if value_index < len(cells) else ""
            if value:
                metadata[label] = value
                index = value_index + 1
                continue
            index += 1
    return metadata


def _find_wechat_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows):
        mapping: dict[str, int] = {}
        for column_index, cell in enumerate(row):
            key = _wechat_header_key(_cell_text(cell))
            if key is not None:
                mapping[key] = column_index
        required = {"time", "sender", "content"}
        if required.issubset(mapping) and ("message_type" in mapping or "sequence" in mapping):
            return row_index, mapping
    raise ValueError("Unsupported WeChat workbook layout.")


def _coalesce_wechat_content(message_type: str, content: str) -> str | None:
    normalized_type = normalize_whitespace(message_type)
    normalized_content = normalize_whitespace(content)
    if normalized_type in WECHAT_TEXT_MESSAGE_TYPES | WECHAT_SYSTEM_MESSAGE_TYPES:
        return normalized_content or None
    if normalized_content.lower() in WECHAT_SKIP_PLACEHOLDERS:
        return None
    if not normalized_content:
        return None
    return f"[{normalized_type}] {normalized_content}"


def parse_wechat_xlsx(raw_bytes: bytes) -> list[NormalizedMessageBase]:
    workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    if not workbook.worksheets:
        return []

    worksheet = workbook.worksheets[0]
    prefix_rows: list[tuple[Any, ...]] = []
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        prefix_rows.append(tuple(row))
        if row_index >= 12:
            break

    header_row_index, header_map = _find_wechat_header_row(prefix_rows)
    metadata = _parse_wechat_metadata(prefix_rows[:header_row_index])
    owner_nickname = metadata.get("\u6635\u79f0") or metadata.get("\u5fae\u4fe1ID") or "\u6211"

    items: list[NormalizedMessageBase] = []
    data_start_row = header_row_index + 2
    for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
        sender = _cell_text(row[header_map["sender"]])
        message_type = _cell_text(row[header_map["message_type"]]) if "message_type" in header_map else ""
        raw_content = _cell_text(row[header_map["content"]])
        if not sender and not raw_content:
            continue

        content = _coalesce_wechat_content(message_type, raw_content)
        if not content:
            continue

        is_self = sender.lower() in WECHAT_SELF_SPEAKERS
        speaker = owner_nickname if is_self and owner_nickname else sender
        role = (
            MessageRole.SYSTEM
            if message_type in WECHAT_SYSTEM_MESSAGE_TYPES
            else MessageRole.ASSISTANT
            if is_self
            else MessageRole.USER
        )

        original_sequence = _cell_text(row[header_map["sequence"]]) if "sequence" in header_map else ""
        items.append(
            NormalizedMessageBase(
                external_id=original_sequence or None,
                speaker=speaker or "\u672a\u77e5",
                role=role,
                content=content,
                occurred_at=parse_datetime(_cell_text(row[header_map["time"]])),
                sequence_index=len(items) + 1,
                metadata={
                    "source_format": WECHAT_SOURCE_FORMAT,
                    "sheet_name": worksheet.title,
                    "message_type": message_type,
                    "original_sender": sender,
                    "is_self": is_self,
                    "conversation_owner": owner_nickname,
                    "export_tool": metadata.get("\u5bfc\u51fa\u5de5\u5177"),
                    "export_version": metadata.get("\u5bfc\u51fa\u7248\u672c"),
                    "platform": metadata.get("\u5e73\u53f0"),
                    "exported_at": metadata.get("\u5bfc\u51fa\u65f6\u95f4"),
                    "original_sequence": original_sequence or None,
                },
            )
        )

    return items


def parse_txt(content: str) -> list[NormalizedMessageBase]:
    items: list[NormalizedMessageBase] = []
    lines = [line.strip() for line in content.splitlines() if line.strip()]
    pattern = re.compile(
        r"^(?:\[(?P<timestamp>[^\]]+)\]\s*)?(?P<speaker>[^:\uff1a]{1,80})[:\uff1a]\s*(?P<content>.+)$"
    )
    for index, line in enumerate(lines, start=1):
        match = pattern.match(line)
        if match:
            speaker = normalize_whitespace(match.group("speaker"))
            body = normalize_whitespace(match.group("content"))
            occurred_at = parse_datetime(match.group("timestamp"))
        else:
            speaker = "Unknown"
            body = normalize_whitespace(line)
            occurred_at = None
        items.append(
            NormalizedMessageBase(
                speaker=speaker,
                role=infer_role(speaker),
                content=body,
                occurred_at=occurred_at,
                sequence_index=index,
                metadata={},
            )
        )
    return items


def parse_csv(content: str) -> list[NormalizedMessageBase]:
    items: list[NormalizedMessageBase] = []
    reader = csv.DictReader(io.StringIO(content))
    for index, row in enumerate(reader, start=1):
        speaker = (
            row.get("speaker")
            or row.get("role")
            or row.get("author")
            or row.get("sender")
            or f"Speaker {index}"
        )
        body = row.get("content") or row.get("message") or row.get("text") or ""
        items.append(
            NormalizedMessageBase(
                external_id=row.get("id") or row.get("message_id"),
                speaker=normalize_whitespace(str(speaker)),
                role=infer_role(str(speaker)),
                content=normalize_whitespace(body),
                occurred_at=parse_datetime(row.get("timestamp") or row.get("created_at") or row.get("time")),
                sequence_index=index,
                metadata={k: v for k, v in row.items() if k not in {"content", "message", "text"}},
            )
        )
    return [item for item in items if item.content]


def parse_json(content: str) -> list[NormalizedMessageBase]:
    payload = json.loads(content)
    if isinstance(payload, dict):
        rows = payload.get("messages") or payload.get("data") or []
    else:
        rows = payload
    items: list[NormalizedMessageBase] = []
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue
        speaker = row.get("speaker") or row.get("role") or row.get("author") or f"Speaker {index}"
        body = row.get("content") or row.get("message") or row.get("text") or ""
        items.append(
            NormalizedMessageBase(
                external_id=str(row.get("id")) if row.get("id") else None,
                speaker=normalize_whitespace(str(speaker)),
                role=infer_role(str(speaker)),
                content=normalize_whitespace(str(body)),
                occurred_at=parse_datetime(row.get("timestamp") or row.get("created_at")),
                sequence_index=index,
                metadata={k: v for k, v in row.items() if k not in {"content", "message", "text"}},
            )
        )
    return [item for item in items if item.content]


def preview_import(file_name: str, raw_bytes: bytes) -> tuple[ImportPreviewResponse, list[NormalizedMessageBase], str]:
    source_type = detect_source_type(file_name)
    content = ""
    if source_type == ImportSourceType.CSV:
        content = raw_bytes.decode("utf-8-sig")
        items = parse_csv(content)
    elif source_type == ImportSourceType.JSON:
        content = raw_bytes.decode("utf-8-sig")
        items = parse_json(content)
    elif source_type == ImportSourceType.XLSX:
        items = parse_wechat_xlsx(raw_bytes)
    else:
        content = raw_bytes.decode("utf-8-sig")
        items = parse_txt(content)
    participants = sorted({item.speaker for item in items})
    source_metadata = _build_source_metadata(items)
    preview = ImportPreviewResponse(
        file_name=file_name,
        source_type=source_type,
        total_messages=len(items),
        detected_participants=participants,
        source_metadata=source_metadata,
        sample_messages=items[:12],
        normalized_messages=items,
    )
    return preview, items, content


def commit_import(db: Session, payload: ImportCommitRequest) -> ImportDetailResponse:
    source_metadata = _build_source_metadata(payload.normalized_messages)

    import_job = ImportJob(
        user_id=payload.user_id,
        file_name=payload.file_name,
        source_type=payload.source_type,
        status=ImportStatus.COMMITTED,
        file_size=payload.file_size,
        raw_text=payload.raw_text,
        raw_payload=payload.raw_payload,
        preview_payload=payload.preview,
        normalized_summary={
            "total_messages": len(payload.normalized_messages),
            "participants": sorted({item.speaker for item in payload.normalized_messages}),
            **source_metadata,
        },
    )
    db.add(import_job)
    db.flush()

    normalized_rows: list[NormalizedMessage] = []
    for item in payload.normalized_messages:
        row = NormalizedMessage(
            import_id=import_job.id,
            external_id=item.external_id,
            speaker=item.speaker,
            role=item.role,
            content=item.content,
            occurred_at=item.occurred_at,
            sequence_index=item.sequence_index,
            details=item.metadata,
        )
        normalized_rows.append(row)
        db.add(row)

    db.commit()
    db.refresh(import_job)
    for row in normalized_rows:
        db.refresh(row)

    return ImportDetailResponse(import_job=import_job, normalized_messages=normalized_rows)
